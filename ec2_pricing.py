
from openpyxl import load_workbook
from collections import namedtuple
import numpy as np
from openpyxl import Workbook

storage = input("(t|f) Include Guestimated EBS?: ")

Instance = namedtuple('Instance', ('name', 'price', 'cpu', 'memory'))
class Server:
    def __init__(self, name, cpu = 0, memory = 0, disks = [], instance = None, poweredOn = True):
        self.name = name
        self.cpu = cpu # count
        self.memory = memory # in mb
        self.disks = disks # array of sizes
        self.instance = instance
        self.poweredOn = poweredOn

def closest(lst, K): 
    lst = np.asarray(lst) 
    idx = (np.abs(lst - K)).argmin() 
    return lst[idx]

wb = load_workbook('JonWoz_CLT-RVTools_export_all_2020-10-21_11.04.43.xlsx')

servers = {}
cpus = wb['vCPU']
rams = wb['vMemory']
drives = wb['vPartition'] #wb['vDisk']
instance_types = wb['InstanceTypes']

# need to make Powerstate aware and understand no volume servers 
# Tom Santuccio -  I do see 32 servers that have no volumes hence no EBS charge. 
# Looks like 11 of those are powered off according to the RVtools output. I presume you could remove all 32 but 
# I’d confirm with the customer. If so, that is another ~53K of EC2 cost you can remove.

for row in cpus.iter_rows(min_row=2, min_col=0, max_col=4):
    name = row[0].value
    servers[name] = Server(name, cpu=int(row[3].value), disks = [], poweredOn = 'poweredOn' == row[1].value)

for row in rams.iter_rows(min_row=2, min_col=0, max_col=4):
    name = row[0].value
    servers[name].memory = int(row[3].value)

for row in drives.iter_rows(min_row=2, min_col=0, max_col=7):
    name = row[0].value
    drive = row[3].value.lower()
    if 'logs' not in drive and 'tmp' not in drive and 'user' not in drive:
        servers[name].disks.append(row[5].value)

# load in instance types, match a type and compute cost
# get default cost for a single type of EBS volue and go with that.
priced_instance_types = {}
instance_prefixes = ('t', 'm6g', 'c6g', 'r6g', 'a1')
cpus = set()
memories = set()
for row in instance_types.iter_rows(min_row=2, min_col=1, max_col=6):
    name = row[0].value
    vcpu = row[1].value
    memory = row[3].value
    if name.startswith(instance_prefixes) and '.' not in memory:
        priced_instance_types['{}:{}'.format(vcpu,memory)] = Instance(name, 
        float(row[5].value.replace('$','').replace('per Hour', '').strip()),
        vcpu,
        memory)  
        cpus.add(vcpu)
        memories.add(int(memory.replace(' GiB','').strip()))     

wb.close

gap = 0
for server in servers.values():
    percent = 1   # this only works well for now with 1 or .5. Have only tested with one RVTools file
    memory = int(closest(list(memories), round(server.memory/1000)) * percent)
    cpu = int(closest(list(cpus), server.cpu) * percent)
    # anomoly correction
    cpu = 1 if cpu == 0 else cpu
    cpu = 4 if cpu == 2 and memory == 32 else cpu
    memory = 2 if memory == 0 else memory
    memory = 8 if cpu == 1 and memory == 16 else memory
    memory = 8 if cpu == 4 and memory == 4 else memory
    memory = 64 if cpu == 8 and memory == 48 else memory
    memory = 16 if cpu == 8 and memory == 8 else memory
    memory = 128 if cpu == 16 and memory == 96 else memory

    for key in priced_instance_types.keys():
        if '{}:{} GiB'.format(cpu,memory) == key:            
            server.instance = priced_instance_types[key]          
    if not server.instance:       
        gap+=1
        if gap < 6:
            print("Could not find an instance type for : {}".format(server.__dict__))
            print('{}:{} GiB'.format(cpu, memory))
if gap > 0:            
    print(gap)
    print(servers['2008r2sp1_dce_customimage'].__dict__)

total = 0
out = Workbook()
filename = 'ec2_estimate.xlsx'
sheet = out.active
sheet.title = 'estimate'
rows = []
rows.append(['server', 'cpu', 'memory', 'volumes', 'total_price', 'instance_type', 'vCPU', 'GiB', ' CPU Price/Hour', 'Annal Instance', 'Annual EBS', 'Annual Snapshot'])
for server in servers.values():
    if (server.poweredOn):
        annaul_instance = server.instance.price * 24 * 365
        annual_ebs = 0
        annaul_snaphot = 0
        if storage.lower() == 't':
            for disk in server.disks:
                annual_ebs = (disk * .1 * 12)/1024
                annaul_snaphot = (disk * .05 * 2 * 12)/1024
        
        server.price = annaul_instance + annual_ebs + annaul_snaphot
        total += server.price
        
        row = []
        rows.append(row)
        row.append(server.name)
        row.append(server.cpu)
        row.append(server.memory)
        row.append(len(server.disks))
        row.append(server.price)
        row.append(server.instance.name)
        row.append(server.instance.cpu)
        row.append(server.instance.memory)
        row.append(server.instance.price)
        row.append(annaul_instance)
        row.append(annual_ebs)
        row.append(annaul_snaphot)
        
print("${:,.2f}".format(total))

for row in rows:
    sheet.append(row)

out.save(filename)

